"""
Lê as abas EMPRESA e FUNCIONARIO de um arquivo .xlsx de admissão.

Localiza o cabeçalho automaticamente pela coluna "Carimbo de data/hora" (âncora que o Google Forms
sempre gera, mesmo que a linha do cabeçalho mude entre abas — ver
references/estrutura-planilha.md). Não fixa o número da linha do cabeçalho no código.

Reescrita em Python (2026-08-11) no lugar da versão PowerShell original — mesma lógica de negócio,
~14x mais rápido (testado: ~60s no PowerShell contra ~4s aqui, no mesmo arquivo real). A versão
PowerShell continua no histórico do projeto se precisar comparar.
"""

from __future__ import annotations

import openpyxl

ANCORA_CABECALHO = "Carimbo de data/hora"
MAX_LINHAS_PARA_PROCURAR_CABECALHO = 5


def _localizar_cabecalho(sheet):
    # Modo read_only do openpyxl usa objetos "EmptyCell" pra células vazias, que não têm .row/.column
    # confiáveis — por isso a posição da coluna vem do índice do enumerate, não do atributo da célula.
    numero_linha = 0
    for linha in sheet.iter_rows(min_row=1, max_row=MAX_LINHAS_PARA_PROCURAR_CABECALHO):
        numero_linha += 1
        for celula in linha:
            if celula.value is not None and ANCORA_CABECALHO in str(celula.value):
                return numero_linha
    raise ValueError(
        f"Não achei a coluna '{ANCORA_CABECALHO}' nas primeiras "
        f"{MAX_LINHAS_PARA_PROCURAR_CABECALHO} linhas da aba '{sheet.title}'. "
        "O formulário pode ter mudado — confira manualmente antes de seguir."
    )


def _ler_aba(sheet, somente_ultimas_n_linhas: int = 0) -> list[dict]:
    linha_cabecalho = _localizar_cabecalho(sheet)

    coluna_para_nome: dict[int, str] = {}
    linhas_cabecalho = list(sheet.iter_rows(min_row=linha_cabecalho, max_row=linha_cabecalho))[0]
    for indice, celula in enumerate(linhas_cabecalho, start=1):
        if celula.value is not None and str(celula.value).strip() != "":
            coluna_para_nome[indice] = str(celula.value)

    resultado = []
    numero_linha_dado = linha_cabecalho
    for linha in sheet.iter_rows(min_row=linha_cabecalho + 1):
        numero_linha_dado += 1
        obj = {"_linha": numero_linha_dado}
        tem_algum_valor = False
        for indice, celula in enumerate(linha, start=1):
            nome_campo = coluna_para_nome.get(indice)
            if not nome_campo:
                continue
            valor = celula.value
            if valor is not None and str(valor).strip() != "":
                tem_algum_valor = True
            obj[nome_campo] = valor
        if tem_algum_valor:
            resultado.append(obj)

    if somente_ultimas_n_linhas > 0 and len(resultado) > somente_ultimas_n_linhas:
        resultado = resultado[-somente_ultimas_n_linhas:]

    return resultado


def ler_planilha_admissao(caminho_arquivo: str, somente_ultimas_n_linhas: int = 600) -> dict:
    """
    Devolve {"empresa": [dict, ...], "funcionario": [dict, ...]}.

    somente_ultimas_n_linhas: otimização, não é filtro de negócio — o corte de data de verdade
    continua sendo aplicado depois, no orquestrador. Padrão generoso (600) cobre quase um ano de
    admissões além de qualquer corte recente, a ~50/mês. Passe 0 para ler a aba inteira.
    """
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True, read_only=True)
    try:
        abas = {ws.title: ws for ws in wb.worksheets}
        for aba_esperada in ("EMPRESA", "FUNCIONARIO"):
            if aba_esperada not in abas:
                raise ValueError(
                    f"A aba '{aba_esperada}' não existe neste arquivo. "
                    f"Abas encontradas: {', '.join(abas.keys())}"
                )

        empresa = _ler_aba(abas["EMPRESA"], somente_ultimas_n_linhas)
        funcionario = _ler_aba(abas["FUNCIONARIO"], somente_ultimas_n_linhas)
        return {"empresa": empresa, "funcionario": funcionario}
    finally:
        wb.close()
